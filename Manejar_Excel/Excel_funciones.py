import openpyxl

class ExcelUtils:
    def __init__(self, driver=None):
        self.driver = driver  # Solo necesario si planeas usar Selenium aquí

    @staticmethod
    def write_data(path, sheet_name, row_num, column_num, data):
        print(f"Escribiendo '{data}' en fila {row_num}, columna {column_num}")
        """Escribe datos en una celda específica de un archivo Excel."""
        workbook = openpyxl.load_workbook(path)
        try:
            sheet = workbook[sheet_name]
            sheet.cell(row=row_num, column=column_num).value = data
            workbook.save(path)
        finally:
            workbook.close()

    @staticmethod
    def get_row_count(path, sheet_name):
        """Devuelve la cantidad de filas utilizadas en una hoja."""
        workbook = openpyxl.load_workbook(path)
        try:
            sheet = workbook[sheet_name]
            return sheet.max_row
        finally:
            workbook.close()

    @staticmethod
    def get_column_count(path, sheet_name):
        """Devuelve la cantidad de columnas utilizadas en una hoja."""
        workbook = openpyxl.load_workbook(path)
        try:
            sheet = workbook[sheet_name]
            return sheet.max_column
        finally:
            workbook.close()

    @staticmethod
    def read_data(path, sheet_name, row_num, column_num):
        workbook = openpyxl.load_workbook(path)
        try:
            sheet = workbook[sheet_name]
            return sheet.cell(row=row_num, column=column_num).value
        finally:
            workbook.close()


